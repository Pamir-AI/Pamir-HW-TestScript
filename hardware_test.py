#!/usr/bin/env python3
"""
Hardware Testing Program for BHV Device
Author: PamirAI
Description: Terminal-based testing program for hardware quality control
"""

import os
import sys
import time
import subprocess
import shutil
import signal
import paramiko
from datetime import datetime
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.hyperlink import Hyperlink
import glob
import threading
import curses
import queue
# Camera features disabled - manual ID entry only
CameraHandler = None
# try:
#     from camera_handler_simple import SimpleCameraHandler as CameraHandler
# except ImportError as e:
#     print(f"Warning: Camera module import failed: {e}")
#     print("QR code scanning disabled.")
#     CameraHandler = None
# except Exception as e:
#     print(f"Error importing camera module: {e}")
#     CameraHandler = None

# Configuration Constants
CM5_IP_ADDRESS = "distiller@192.168.0.30"
CM5_PASSWORD = "one"
EXCEL_FILE = "hardware_test_results.xlsx"
RGB_LED_TEST_PATH = "/opt/distiller-cm5-sdk/src/distiller_cm5_sdk/hardware/sam/led_interactive_demo.py"
RGB_LED_NUM_ENTERS = 10  # Number of Enter presses needed for the RGB LED interactive program

# UF2 Configuration (from upload.py)
UF2_DIRECTORY = "ULP"
VOLUME_PATHS = ["/Volumes/RPI-RP2 1", "/Volumes/RPI-RP2"]
UART_PORT_PATTERN = "/dev/tty.usb*"
FLASH_NUKE_UF2 = "flash_nuke.uf2"
MICROPYTHON_UF2 = "RPI_PICO-20240222-v1.22.2.uf2"

# Files to upload (from BHV folder)
PYTHON_FILES = [
    "bin/loading1.bin",
    "bin/loading2.bin",
    "eink_driver_sam.py",
    "pamir_uart_protocols.py",
    "neopixel_controller.py",
    "power_manager.py",
    "battery.py",
    "debug_handler.py",
    "uart_handler.py",
    "threaded_task_manager.py",
    "main.py",
]

# Test ID definitions
TEST_IDS = {
    'T01': 'Firmware Upload',
    'T02': 'CM5 LED Visual Check',
    'T03': 'RGB LED Visual Check',
    'T04': 'E-ink Display Refresh',
    'T05': 'UI Appears',
    'T06': 'Button Response',
    'T07': 'Voice Transcribed',
    'T09': 'USB MicroPython Detection',
    'T10': 'USB Hub Detection',
    'T11': 'USB Media Detection',
    'T12': 'RGB LED SSH Test',
    'T13': 'SD Card Detection',
    'T14': 'Camera Detection'
}

# Test result structure
class TestResult:
    def __init__(self, device_id, tests_to_run, version=None, manufacture_id=None):
        self.device_id = device_id
        self.version = version  # From QR code
        self.manufacture_id = manufacture_id  # From QR code
        self.timestamp = datetime.now()
        self.test_results = {}  # Store results by test ID
        self.test_durations = {}  # Store duration in seconds by test ID
        self.tests_to_run = tests_to_run  # Set of test IDs to run
        self.overall_pass = False
        self.notes = ""
        # Create logs directory if it doesn't exist
        os.makedirs("logs", exist_ok=True)
        self.log_file = f"logs/device_{device_id}_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        self.video_file = None  # Path to recorded video
        
        # Initialize all tests
        for test_id in TEST_IDS:
            if test_id in tests_to_run:
                self.test_results[test_id] = None
                self.test_durations[test_id] = None
            else:
                self.test_results[test_id] = 'SKIPPED'
                self.test_durations[test_id] = 0
    
    def set_test_result(self, test_id, result, details="", duration=None):
        """Set individual test result and log to file"""
        self.test_results[test_id] = result
        if duration is not None:
            self.test_durations[test_id] = duration
        
        status = 'PASS' if result else 'FAIL' if result is not None else 'SKIPPED'
        log_entry = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {test_id}: {TEST_IDS[test_id]} - {status}"
        if duration is not None:
            log_entry += f" - Duration: {duration}s"
        if details:
            log_entry += f" - {details}"
        
        # Write to log file
        with open(self.log_file, 'a') as f:
            f.write(log_entry + "\n")
        
        print(log_entry)

def clear_screen():
    """Clear terminal screen"""
    os.system('cls' if os.name == 'nt' else 'clear')

def print_header(text):
    """Print formatted header"""
    print("\n" + "="*60)
    print(f" {text} ".center(60))
    print("="*60 + "\n")

def get_yes_no_input(prompt, allow_continue_check=True):
    """Get yes/no input from user with timing"""
    start_time = time.time()
    while True:
        response = input(f"{prompt} (y/n): ").strip().lower()
        if response in ['y', 'yes']:
            return True, int(time.time() - start_time)
        elif response in ['n', 'no']:
            if allow_continue_check:
                continue_test = input("Test failed. Do you want to continue testing? (y/n): ").strip().lower()
                if continue_test not in ['y', 'yes']:
                    return None, int(time.time() - start_time)  # Signal to restart
            return False, int(time.time() - start_time)
        else:
            print("Please enter 'y' for yes or 'n' for no.")

def wait_for_enter(prompt):
    """Wait for user to press Enter"""
    input(f"{prompt} Press Enter when ready: ")

# Upload functionality adapted from upload.py
def find_rp2_volume():
    """Find the active RPI-RP2 volume"""
    for volume_path in VOLUME_PATHS:
        if os.path.exists(volume_path):
            try:
                files = os.listdir(volume_path)
                if "INFO_UF2.TXT" in files:
                    return volume_path
            except (PermissionError, OSError):
                continue
    return None

def wait_for_rp2_device(timeout=60):
    """Wait for RPI-RP2 device to appear"""
    print("Waiting for RPI-RP2 device...")
    start_time = time.time()
    
    while time.time() - start_time < timeout:
        volume_path = find_rp2_volume()
        if volume_path:
            print(f"RPI-RP2 device found at: {volume_path}")
            return volume_path
        time.sleep(1)
    
    print(f"Timeout: RPI-RP2 device not found after {timeout} seconds")
    return None

def wait_for_rp2_disappear(timeout=30):
    """Wait for RPI-RP2 device to disappear"""
    print("Waiting for RPI-RP2 device to disappear...")
    start_time = time.time()
    
    while time.time() - start_time < timeout:
        if not find_rp2_volume():
            print("RPI-RP2 device disappeared")
            return True
        time.sleep(1)
    
    print(f"Timeout: RPI-RP2 device still present after {timeout} seconds")
    return False

def flash_uf2_file(uf2_filename, description):
    """Flash a UF2 file to the RP2040"""
    uf2_path = os.path.join(UF2_DIRECTORY, uf2_filename)
    
    if not os.path.exists(uf2_path):
        print(f"Error: {uf2_filename} not found in {UF2_DIRECTORY}")
        return False
    
    volume_path = find_rp2_volume()
    if not volume_path:
        print("No RPI-RP2 volume found")
        return False
    
    print(f"Copying {description}...")
    
    try:
        dest_path = os.path.join(volume_path, uf2_filename)
        shutil.copy2(uf2_path, dest_path)
        print(f"{description} copied successfully")
        return True
    except Exception as e:
        # For MicroPython firmware, the error is expected as device disappears quickly
        if "RPI_PICO" in uf2_filename and "Invalid argument" in str(e):
            print(f"{description} copied successfully (device auto-ejected)")
            return True
        print(f"Copy failed: {e}")
        return False

def select_tests_to_run():
    """Interactive menu to select which tests to run"""
    def menu(stdscr):
        curses.curs_set(0)  # Hide cursor
        stdscr.clear()
        
        # Initialize all tests as selected
        selected_tests = set(TEST_IDS.keys())
        current_row = 0
        
        while True:
            stdscr.clear()
            height, width = stdscr.getmaxyx()
            
            # Title
            title = "Select Tests to Run (Space to toggle, Enter to confirm, A to select all, N to select none)"
            stdscr.addstr(0, (width - len(title)) // 2, title, curses.A_BOLD)
            
            # Display tests
            for idx, (test_id, test_name) in enumerate(TEST_IDS.items()):
                y = idx + 2
                if y >= height - 2:
                    break
                    
                # Checkbox
                checkbox = "[X]" if test_id in selected_tests else "[ ]"
                
                # Highlight current row
                if idx == current_row:
                    stdscr.attron(curses.A_REVERSE)
                    
                stdscr.addstr(y, 2, f"{checkbox} {test_id}: {test_name}")
                
                if idx == current_row:
                    stdscr.attroff(curses.A_REVERSE)
            
            # Instructions at bottom
            stdscr.addstr(height - 1, 2, "Use ↑/↓ to navigate, Space to toggle, Enter to confirm")
            
            # Get user input
            key = stdscr.getch()
            
            if key == curses.KEY_UP and current_row > 0:
                current_row -= 1
            elif key == curses.KEY_DOWN and current_row < len(TEST_IDS) - 1:
                current_row += 1
            elif key == ord(' '):  # Space to toggle
                test_id = list(TEST_IDS.keys())[current_row]
                if test_id in selected_tests:
                    selected_tests.remove(test_id)
                else:
                    selected_tests.add(test_id)
            elif key == ord('a') or key == ord('A'):  # Select all
                selected_tests = set(TEST_IDS.keys())
            elif key == ord('n') or key == ord('N'):  # Select none
                selected_tests = set()
            elif key == ord('\n'):  # Enter to confirm
                return selected_tests
            elif key == 27:  # ESC to exit
                return set(TEST_IDS.keys())  # Return all selected by default
    
    try:
        return curses.wrapper(menu)
    except:
        # Fallback if curses fails
        print("\nSelect tests to run (interactive menu failed, using text mode):")
        print("Enter test IDs to skip (comma-separated, or press Enter to run all):")
        for test_id, test_name in TEST_IDS.items():
            print(f"  {test_id}: {test_name}")
        
        skip_input = input("\nTests to skip (e.g., T01,T05,T10): ").strip()
        if skip_input:
            skip_tests = set(t.strip().upper() for t in skip_input.split(','))
            return set(TEST_IDS.keys()) - skip_tests
        return set(TEST_IDS.keys())

def upload_firmware_wipe(test_result):
    """Upload firmware with --wipe functionality"""
    print_header("Uploading Firmware (Wipe Mode)")
    start_time = time.time()
    
    # Wait for device in bootloader mode
    volume_path = wait_for_rp2_device()
    if not volume_path:
        duration = int(time.time() - start_time)
        test_result.set_test_result('T01', False, "No RPI-RP2 volume found", duration)
        return False
    
    # Flash nuke first
    if not flash_uf2_file(FLASH_NUKE_UF2, "flash nuke"):
        duration = int(time.time() - start_time)
        test_result.set_test_result('T01', False, "Flash nuke failed", duration)
        return False
    
    # Wait for device to disappear and reappear
    if not wait_for_rp2_disappear():
        duration = int(time.time() - start_time)
        test_result.set_test_result('T01', False, "Device did not disappear after flash nuke", duration)
        return False
    
    print("Waiting for device to reappear...")
    time.sleep(5)
    
    volume_path = wait_for_rp2_device()
    if not volume_path:
        duration = int(time.time() - start_time)
        test_result.set_test_result('T01', False, "Device did not reappear after flash nuke", duration)
        return False
    
    # Flash MicroPython firmware
    if not flash_uf2_file(MICROPYTHON_UF2, "MicroPython firmware"):
        duration = int(time.time() - start_time)
        test_result.set_test_result('T01', False, "MicroPython firmware flash failed", duration)
        return False
    
    # Wait for device to disappear and initialize
    wait_for_rp2_disappear()
    print("Waiting for device to initialize...")
    time.sleep(3)
    
    # Find UART port
    ports = glob.glob(UART_PORT_PATTERN)
    if not ports:
        print(f"No UART ports found matching pattern: {UART_PORT_PATTERN}")
        duration = int(time.time() - start_time)
        test_result.set_test_result('T01', False, "No UART port found", duration)
        return False
    
    port = ports[0]
    print(f"Using port: {port}")
    
    # Upload Python files
    bhv_dir = Path(__file__).parent / "BHV"
    total_files = len(PYTHON_FILES)
    
    for i, filename in enumerate(PYTHON_FILES):
        file_path = bhv_dir / filename
        
        if not file_path.exists():
            print(f"Warning: {filename} not found, skipping...")
            continue
        
        print(f"Uploading {filename} ({i+1}/{total_files})")
        
        try:
            cmd = ["ampy", "--port", port, "put", str(file_path)]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            
            if result.returncode != 0:
                print(f"Error uploading {filename}: {result.stderr}")
                duration = int(time.time() - start_time)
                test_result.set_test_result('T01', False, f"Failed to upload {filename}", duration)
                return False
            
        except Exception as e:
            print(f"Error uploading {filename}: {e}")
            duration = int(time.time() - start_time)
            test_result.set_test_result('T01', False, f"Exception uploading {filename}: {e}", duration)
            return False
    
    print("\nFirmware upload completed successfully!")
    duration = int(time.time() - start_time)
    test_result.set_test_result('T01', True, "All files uploaded successfully", duration)
    return True

def ssh_execute_command(ssh_client, command, send_enter=False):
    """Execute command via SSH and return output"""
    try:
        stdin, stdout, stderr = ssh_client.exec_command(command)
        
        # If command needs Enter key, send it
        if send_enter:
            stdin.write('\n')
            stdin.flush()
        
        output = stdout.read().decode('utf-8')
        error = stderr.read().decode('utf-8')
        return output, error
    except Exception as e:
        print(f"SSH command error: {e}")
        return "", f"SSH_ERROR: {str(e)}"

def ssh_execute_interactive(ssh_client, command, num_enters=1, enter_delay=0.5):
    """Execute interactive command via SSH that requires multiple Enter presses"""
    try:
        # Use invoke_shell for interactive commands
        channel = ssh_client.invoke_shell()
        time.sleep(0.5)  # Wait for shell to be ready
        
        # Send the command
        channel.send(command + '\n')
        time.sleep(1)  # Wait for command to start
        
        # Send multiple Enter presses
        for i in range(num_enters):
            channel.send('\n')
            time.sleep(enter_delay)
        
        # Collect output
        output = ""
        while channel.recv_ready():
            output += channel.recv(1024).decode('utf-8')
        
        channel.close()
        return output, ""
    except Exception as e:
        print(f"SSH interactive command error: {e}")
        return "", str(e)

def reconnect_ssh(test_result, max_attempts=5):
    """Reconnect SSH with retry logic"""
    for attempt in range(max_attempts):
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            
            user, host = CM5_IP_ADDRESS.split('@')
            
            print(f"\nSSH reconnection attempt {attempt + 1}/{max_attempts}...")
            print(f"Connecting to {CM5_IP_ADDRESS}...")
            ssh.connect(host, username=user, password=CM5_PASSWORD, timeout=30)
            print("Reconnected successfully!")
            return ssh
            
        except Exception as e:
            print(f"SSH reconnection failed: {e}")
            if attempt < max_attempts - 1:
                wait_for_enter("\nPress Enter to retry SSH connection")
            else:
                print("\nAll SSH reconnection attempts failed.")
                return None
    return None

def perform_ssh_tests(test_result):
    """Perform automated tests via SSH with retry logic"""
    print_header("Performing SSH Tests")
    
    # Initial SSH connection
    ssh = reconnect_ssh(test_result)
    if not ssh:
        test_result.notes += " Initial SSH connection failed after 5 attempts."
        # Mark all SSH tests as failed
        ssh_test_ids = ['T09', 'T10', 'T11', 'T12', 'T13', 'T14']
        for test_id in ssh_test_ids:
            if test_id in test_result.tests_to_run:
                test_result.set_test_result(test_id, False, "SSH connection failed", 0)
        return False
    
    try:
        
        # Test 1: Check USB devices
        print("\nChecking USB devices...")
        output, error = ssh_execute_command(ssh, "lsusb")
        
        # Check if SSH connection dropped and reconnect if needed
        if "SSH_ERROR" in error:
            print("SSH connection lost. Attempting to reconnect...")
            if ssh:
                ssh.close()
            ssh = reconnect_ssh(test_result)
            if ssh:
                output, error = ssh_execute_command(ssh, "lsusb")
            else:
                print("Failed to reconnect SSH for USB tests")
        
        # T09: USB MicroPython
        if 'T09' in test_result.tests_to_run:
            start_time = time.time()
            time.sleep(1)
            usb_micropython = "MicroPython Board in FS mode" in output
            duration = int(time.time() - start_time)
            test_result.set_test_result('T09', usb_micropython, f"lsusb output: {output[:100]}...", duration)
        
        # T10: USB Hub
        if 'T10' in test_result.tests_to_run:
            start_time = time.time()
            usb_hub = "QinHeng Electronics USB HUB" in output
            duration = int(time.time() - start_time)
            test_result.set_test_result('T10', usb_hub, "", duration)
        
        # T11: USB Media
        if 'T11' in test_result.tests_to_run:
            start_time = time.time()
            usb_media = "Microchip Technology, Inc. (formerly SMSC) Ultra Fast Media" in output
            duration = int(time.time() - start_time)
            test_result.set_test_result('T11', usb_media, "", duration)
        
        # Test 2: RGB LED test
        if 'T12' in test_result.tests_to_run:
            print(f"\nRunning RGB LED test...")
            print(f"Executing: python3 {RGB_LED_TEST_PATH}")
            start_time = time.time()
            
            # First check if the file exists
            check_output, check_error = ssh_execute_command(ssh, f"test -f {RGB_LED_TEST_PATH} && echo 'EXISTS' || echo 'NOT_FOUND'")
            
            if 'NOT_FOUND' in check_output:
                print(f"Error: RGB LED test script not found at {RGB_LED_TEST_PATH}")
                test_result.set_test_result('T12', False, "Script not found", int(time.time() - start_time))
            else:
                print("RGB LED test program is running...")
                print("The program will cycle through LED colors. Please observe the LED.")
                
                # Run the interactive LED program with multiple Enter presses
                output, error = ssh_execute_interactive(ssh, f"python3 {RGB_LED_TEST_PATH}", num_enters=RGB_LED_NUM_ENTERS, enter_delay=0.5)
                
                # Check if we need to reconnect
                if "SSH_ERROR" in error:
                    print("SSH connection lost during LED test. Attempting to reconnect and retry...")
                    if ssh:
                        ssh.close()
                    ssh = reconnect_ssh(test_result)
                    if ssh:
                        output, error = ssh_execute_interactive(ssh, f"python3 {RGB_LED_TEST_PATH}", num_enters=RGB_LED_NUM_ENTERS, enter_delay=0.5)
                
                # Give some time for the program to complete
                time.sleep(2)
                
                # Ask user if they saw the LED
                rgb_led_result, question_duration = get_yes_no_input("Did you see the RGB LED cycle through different colors?", allow_continue_check=False)
                
                # Total duration includes script execution + user response time
                total_duration = int(time.time() - start_time)
                test_result.set_test_result('T12', rgb_led_result, "", total_duration)
        
        
        
        # Test 3: Check SD card
        if 'T13' in test_result.tests_to_run:
            print("\nChecking SD card...")
            start_time = time.time()
            time.sleep(1)
            output, error = ssh_execute_command(ssh, "lsblk")
            
            # Reconnect if needed
            if "SSH_ERROR" in error:
                print("SSH connection lost. Attempting to reconnect...")
                if ssh:
                    ssh.close()
                ssh = reconnect_ssh(test_result)
                if ssh:
                    output, error = ssh_execute_command(ssh, "lsblk")
                else:
                    test_result.set_test_result('T13', False, "SSH reconnection failed", int(time.time() - start_time))
            
            if ssh and "SSH_ERROR" not in error:  # Only process if we have a valid connection
                sd_card_result = "sda1" in output
                duration = int(time.time() - start_time)
                test_result.set_test_result('T13', sd_card_result, f"lsblk output: {output[:100]}...", duration)
        
        # Test 4: Check camera
        if 'T14' in test_result.tests_to_run:
            print("\nChecking camera...")
            start_time = time.time()
            output, error = ssh_execute_command(ssh, "libcamera-hello")
            
            # Reconnect if needed
            if "SSH_ERROR" in error:
                print("SSH connection lost. Attempting to reconnect...")
                if ssh:
                    ssh.close()
                ssh = reconnect_ssh(test_result)
                if ssh:
                    output, error = ssh_execute_command(ssh, "libcamera-hello")
                else:
                    test_result.set_test_result('T14', False, "SSH reconnection failed", int(time.time() - start_time))
            
            if ssh and "SSH_ERROR" not in error:  # Only process if we have a valid connection
                camera_result = "ERROR: *** no cameras available ***" not in error
                duration = int(time.time() - start_time)
                test_result.set_test_result('T14', camera_result, error[:100] if error else "Camera OK", duration)
        
        if ssh:
            ssh.close()
        return True
        
    except Exception as e:
        print(f"SSH error during tests: {e}")
        test_result.notes += f" SSH error: {str(e)}"
        if ssh:
            ssh.close()
        return True  # Return True to continue with other tests

def save_to_excel(test_result):
    """Save test results to Excel file"""
    print_header("Saving Results")
    
    # Calculate overall pass/fail
    test_results_values = [v for v in test_result.test_results.values() if v is not None]
    test_result.overall_pass = all(test_results_values) if test_results_values else False
    
    # Create DataFrame with all test IDs
    data = {
        'Device ID': test_result.device_id,
        'Version': test_result.version if test_result.version else '',
        'Manufacture ID': test_result.manufacture_id if test_result.manufacture_id else '',
        'Timestamp': test_result.timestamp,
        'Log File': test_result.log_file,
        'Video File': test_result.video_file if test_result.video_file else ''
    }
    
    # First add all test results
    for test_id, test_name in TEST_IDS.items():
        result = test_result.test_results.get(test_id)
        
        if result == 'SKIPPED':
            data[f'{test_id}: {test_name}'] = 'SKIPPED'
        elif result is None:
            data[f'{test_id}: {test_name}'] = 'NOT RUN'
        else:
            data[f'{test_id}: {test_name}'] = 'PASS' if result else 'FAIL'
    
    # Add failed tests summary and overall pass
    failed_tests = [f"{tid}({TEST_IDS[tid]})" for tid, result in test_result.test_results.items() if result is False]
    data['Failed Tests'] = ', '.join(failed_tests) if failed_tests else 'None'
    data['Overall Pass'] = test_result.overall_pass
    data['Notes'] = test_result.notes
    
    # Then add all durations at the end
    for test_id, test_name in TEST_IDS.items():
        result = test_result.test_results.get(test_id)
        duration = test_result.test_durations.get(test_id, 0)
        
        if result == 'SKIPPED' or result is None:
            data[f'{test_id} Duration (s)'] = 0
        else:
            data[f'{test_id} Duration (s)'] = duration if duration is not None else 0
    
    # Check if Excel file exists
    if os.path.exists(EXCEL_FILE):
        # Load existing data
        df = pd.read_excel(EXCEL_FILE)
        # Append new data
        df = pd.concat([df, pd.DataFrame([data])], ignore_index=True)
    else:
        # Create new DataFrame
        df = pd.DataFrame([data])
    
    # Save to Excel with formatting
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Test Results', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Test Results']
        
        # Format headers
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Color code pass/fail cells
        for row in range(2, len(df) + 2):
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if col_name.startswith('T') and ':' in col_name:  # Test result columns
                    if cell.value == 'PASS':
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif cell.value == 'FAIL':
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    elif cell.value == 'NOT RUN':
                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    elif cell.value == 'SKIPPED':
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                elif col_name == 'Overall Pass':
                    if cell.value:
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif col_name == 'Log File' and cell.value:
                    # Make log file a clickable hyperlink with absolute path
                    abs_path = os.path.abspath(cell.value)
                    cell.hyperlink = abs_path
                    cell.font = Font(color="0000FF", underline="single")
                    cell.value = os.path.basename(cell.value)  # Show only filename in cell
                elif col_name == 'Video File' and cell.value:
                    # Make video file a clickable hyperlink with absolute path
                    abs_path = os.path.abspath(cell.value)
                    cell.hyperlink = abs_path
                    cell.font = Font(color="0000FF", underline="single")
                    cell.value = os.path.basename(cell.value)  # Show only filename in cell
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"Results saved to {EXCEL_FILE}")
    print(f"Test Result: {'PASS' if test_result.overall_pass else 'FAIL'}")

def setup_ctrl_x_handler():
    """Setup Ctrl+X handler for graceful exit"""
    def signal_handler(sig, frame):
        print("\n\nCtrl+X pressed. Exiting program...")
        sys.exit(0)
    
    # Note: On Unix systems, we can't directly capture Ctrl+X, so we use Ctrl+C
    signal.signal(signal.SIGINT, signal_handler)

def get_device_id_with_qr(camera):
    """Get device ID from manual input"""
    device_id = None
    version = None
    manufacture_id = None
    
    print("\n" + "="*60)
    print("DEVICE ID INPUT")
    print("="*60)
    
    # Camera features disabled - manual input only
    # if camera:
    #     print("\nStarting camera preview for QR code scanning...")
    #     print("This will open a camera window.")
    #     print("Press 'q' in the camera window when done.\n")
    #     
    #     # Run camera preview in main thread
    #     camera.run_camera_preview()
    #     
    #     # Check if QR was scanned
    #     qr_data = camera.get_qr_data(timeout=0.1)
    #     if qr_data:
    #         device_id = int(qr_data['device_id'])
    #         version = qr_data['version']
    #         manufacture_id = qr_data['manufacture_id']
    #         print(f"\nUsing QR scanned data:")
    #         print(f"Device ID: {device_id}")
    #         print(f"Version: {version}")
    #         print(f"Manufacture ID: {manufacture_id}")
    #         return device_id, version, manufacture_id
    
    # Manual input only
    while device_id is None:
        id_input = input("Enter device ID (integer): ")
        if id_input.isdigit():
            device_id = int(id_input)
            print(f"Device ID = {device_id}")
            break
        else:
            print("Invalid input. Please enter a number.")
    
    return device_id, version, manufacture_id

def main():
    """Main testing program"""
    setup_ctrl_x_handler()
    
    # Initialize camera variables
    camera = None
    camera_started = False
    
    try:
        while True:  # Main loop for restarting tests
            clear_screen()
            print_header("BHV Hardware Testing Program")
            print("This program will guide you through testing a hardware unit.")
            print("Please follow all instructions carefully.")
            print("\n[Hint: Press Ctrl+C to quit the program at any time]\n")
            
            # Camera features disabled
            # if not camera_started and CameraHandler is not None:
            #     try:
            #         print("\n" + "="*60)
            #         print("CAMERA SETUP")
            #         print("="*60 + "\n")
            #         camera = CameraHandler()
            #         camera_started = True
            #         print("\nCamera initialized. Preview will start when needed.")
            #     except Exception as e:
            #         print(f"\nCamera initialization error: {e}")
            #         print("Manual device ID entry only.")
            # elif CameraHandler is None:
            #     print("\nCamera module not available. Manual device ID entry only.")
            
            # Select tests to run
            print("\nSelect tests to run...")
            tests_to_run = select_tests_to_run()
            print(f"\nSelected {len(tests_to_run)} tests to run")
            
            # Get device ID (manual or QR code)
            device_id, version, manufacture_id = get_device_id_with_qr(camera if camera_started else None)
            
            test_result = TestResult(device_id, tests_to_run, version, manufacture_id)
            restart_test = False
        
            try:
                # Step 1: Firmware upload
                if 'T01' in test_result.tests_to_run:
                    wait_for_enter("Put the board into bootloader mode and connect to computer.")
                    firmware_success = upload_firmware_wipe(test_result)
                    
                    if not firmware_success:
                        print("\nFirmware upload failed! Cannot continue testing.")
                        test_result.notes = "Firmware upload failed"
                        save_to_excel(test_result)
                        continue
                else:
                    print("\nSkipping firmware upload (T01)")
                
                # Step 2: Hardware assembly - only if we have tests that need hardware
                hardware_tests = {'T02', 'T03', 'T04', 'T05', 'T06', 'T07'}
                if any(test in test_result.tests_to_run for test in hardware_tests):
                    wait_for_enter("\nUnplug the board and insert: battery, camera module, e-ink display, SD card, and CM5.")
                
                # Step 3: Visual checks - only if we have visual tests
                visual_tests = {'T02', 'T03', 'T04'}
                if any(test in test_result.tests_to_run for test in visual_tests):
                    print_header("Visual Checks")
                
                # T02: CM5 LED
                if 'T02' in test_result.tests_to_run:
                    result, duration = get_yes_no_input("Does the CM5 LED light up?")
                    if result is None:
                        restart_test = True
                        save_to_excel(test_result)
                        continue
                    test_result.set_test_result('T02', result, "", duration)
            
                # T03: RGB LED Visual
                if 'T03' in test_result.tests_to_run:
                    result, duration = get_yes_no_input("Does the RGB LED light up?")
                    if result is None:
                        restart_test = True
                        save_to_excel(test_result)
                        continue
                    test_result.set_test_result('T03', result, "", duration)
                
                # T04: E-ink Display
                if 'T04' in test_result.tests_to_run:
                    result, duration = get_yes_no_input("Does the e-ink bootup screen appear?")
                    if result is None:
                        restart_test = True
                        save_to_excel(test_result)
                        continue
                    test_result.set_test_result('T04', result, "", duration)
                
                # Step 4: CM5 UI tests - only if we have UI tests
                ui_tests = {'T05', 'T06', 'T07'}
                if any(test in test_result.tests_to_run for test in ui_tests):
                    print_header("CM5 UI Tests")
                
                # T05: UI Appears
                ui_appears = False
                if 'T05' in test_result.tests_to_run:
                    result, duration = get_yes_no_input("Wait for WIFI UI to show up. Does the WIFI UI appear?")
                    if result is None:
                        restart_test = True
                        save_to_excel(test_result)
                        continue
                    ui_appears = result
                    test_result.set_test_result('T05', result, "", duration)
                
                # Continue UI tests only if UI appears or T05 was skipped
                if ui_appears or 'T05' not in test_result.tests_to_run:
                    # T06: Button Response
                    if 'T06' in test_result.tests_to_run:
                        result, duration = get_yes_no_input("Press a button. Does the button show on dmesg?")
                        if result is None:
                            restart_test = True
                            save_to_excel(test_result)
                            continue
                        test_result.set_test_result('T06', result, "", duration)
                    
                    # # Only show instructions if we have voice tests
                    # if 'T07' in test_result.tests_to_run:
                    #     print("\nInstructions:")
                    #     print("1. Go to 'select server' button and select 'medical assistant server'")
                    #     print("2. Wait for cache restoration")
                    #     print("3. If you see 'failed to connect', reload the page")
                    #     wait_for_enter("")
                    
                    # T07: Voice Transcribed
                    voice_transcribed = False
                    if 'T07' in test_result.tests_to_run:
                        print("Use Piper to perform audio record and speak a test message.")
                        result, duration = get_yes_no_input("Was your voice successfully transcribed?")
                        if result is None:
                            restart_test = True
                            save_to_excel(test_result)
                            continue
                        voice_transcribed = result
                        test_result.set_test_result('T07', result, "", duration)
                    
                
                # Step 5: SSH tests - only if we have SSH tests
                ssh_tests = {'T09', 'T10', 'T11', 'T12', 'T13', 'T14'}
                if any(test in test_result.tests_to_run for test in ssh_tests):
                    if perform_ssh_tests(test_result):
                        print("\nSSH tests completed successfully!")
                    else:
                        print("\nSSH tests failed or incomplete")
                        test_result.notes += " SSH tests incomplete."
                
                # Camera recording disabled
                # if camera_started and camera and camera.recording:
                #     print("\nSaving test video...")
                #     video_file = camera.stop_recording_and_save(test_result.device_id, test_result.timestamp.strftime('%Y%m%d_%H%M%S'))
                #     if video_file:
                #         test_result.video_file = video_file
                
                # Save results
                save_to_excel(test_result)
                
                print_header("Testing Complete")
                print(f"Device ID: {test_result.device_id}")
                if test_result.version:
                    print(f"Version: {test_result.version}")
                    print(f"Manufacture ID: {test_result.manufacture_id}")
                print(f"Overall Result: {'PASS' if test_result.overall_pass else 'FAIL'}")
                
                # Ask if user wants to shutdown CM5
                shutdown_response = input("\nPress Enter to safely shutdown the CM5, or 'n' to skip: ").strip().lower()
                if shutdown_response != 'n':
                    print("Shutting down CM5...")
                    try:
                        # Try to connect if we don't have an active connection
                        ssh = reconnect_ssh(test_result, max_attempts=3)
                        if ssh:
                            output, error = ssh_execute_command(ssh, "sudo shutdown now")
                            print("Shutdown command sent. The CM5 will power off in a few seconds.")
                            ssh.close()
                        else:
                            print("Could not connect to CM5 for shutdown.")
                    except Exception as e:
                        print(f"Error during shutdown: {e}")
                
                # Ask if user wants to test another device
                another = input("\nDo you want to test another device? (y/n): ").strip().lower()
                if another not in ['y', 'yes']:
                    break
                    
            except KeyboardInterrupt:
                print("\n\nProgram terminated by user")
                if 'test_result' in locals():
                    test_result.notes += " Test interrupted."
                    save_to_excel(test_result)
                sys.exit(0)
            except Exception as e:
                print(f"\n\nError occurred: {e}")
                if 'test_result' in locals():
                    test_result.notes += f" Error: {str(e)}"
                    save_to_excel(test_result)
                input("Press Enter to restart testing...")
    
    finally:
        # Cleanup camera
        if camera_started and camera:
            camera.stop_preview_and_camera()

if __name__ == "__main__":
    main()